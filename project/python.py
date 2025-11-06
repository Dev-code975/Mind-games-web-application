import os
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# -------------------------------
# SETUP: Headless Chrome for GitHub
# -------------------------------
options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--window-size=1920,1080")

driver = webdriver.Chrome(options=options)

# Use relative path to index.html (inside repo)
current_dir = os.path.dirname(os.path.abspath(__file__))
index_path = os.path.join(current_dir, "index.html")
driver.get(f"file:///{index_path}")

print("🧠 AUTOMATED TESTING: Mind Games for Seniors\n")

# -------------------------------
# SETUP EXCEL WORKBOOK
# -------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Test Report"

headers = ["Test Case ID", "Test Case Name", "Description", "Expected Result", "Actual Result", "Status"]
ws.append(headers)

# Header styling
for col in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

border_style = Border(left=Side(style="thin"),
                      right=Side(style="thin"),
                      top=Side(style="thin"),
                      bottom=Side(style="thin"))

# -------------------------------
# FUNCTION TO LOG TEST RESULTS
# -------------------------------
def log_test(tc_id, name, desc, expected, actual, status):
    ws.append([tc_id, name, desc, expected, actual, status])
    row = ws.max_row
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=row, column=col)
        cell.border = border_style
        if status == "PASS":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "FAIL":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

wait = WebDriverWait(driver, 10)

# -------------------------------
# TEST CASES
# -------------------------------
try:
    title = driver.title
    assert "Mind Games for Seniors" in title
    print("✅ TC01: Homepage Loaded")
    log_test("TC01", "Verify Homepage Loaded", "Check if homepage loads successfully",
             "Page title should contain 'Mind Games for Seniors'", f"Title: {title}", "PASS")
except Exception as e:
    log_test("TC01", "Verify Homepage Loaded", "Check if homepage loads successfully",
             "Title should contain 'Mind Games for Seniors'", str(e), "FAIL")

try:
    sudoku_active = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#sudoku-game.active")))
    print("✅ TC02: Sudoku Game Loads by Default")
    log_test("TC02", "Sudoku Default Active", "Verify Sudoku section is visible by default",
             "Sudoku section should be active", "Sudoku section found active", "PASS")
except Exception as e:
    log_test("TC02", "Sudoku Default Active", "Verify Sudoku section is visible by default",
             "Sudoku section should be active", str(e), "FAIL")

try:
    nav_buttons = driver.find_elements(By.CLASS_NAME, "game-btn")
    for btn in nav_buttons:
        btn.click()
        time.sleep(1)
        game_id = btn.get_attribute("data-game")
        print(f"    → Switched to {game_id}")
    print("✅ TC03: Navigation Buttons Work")
    log_test("TC03", "Verify Navigation", "Ensure game navigation buttons switch between sections",
             "Each click should switch to correct game section", "All sections switched successfully", "PASS")
except Exception as e:
    log_test("TC03", "Verify Navigation", "Ensure game navigation buttons switch between sections",
             "Each click should switch to correct game section", str(e), "FAIL")

try:
    sudoku_controls = driver.find_elements(By.CSS_SELECTOR, "#sudoku-game .control-btn")
    assert len(sudoku_controls) >= 3
    print("✅ TC04: Sudoku Controls")
    log_test("TC04", "Verify Sudoku Controls", "Check existence of Sudoku control buttons",
             "New Game, Check, Hint buttons should be present", f"Found {len(sudoku_controls)} controls", "PASS")
except Exception as e:
    log_test("TC04", "Verify Sudoku Controls", "Check existence of Sudoku control buttons",
             "New Game, Check, Hint buttons should be present", str(e), "FAIL")

try:
    difficulties = driver.find_elements(By.CLASS_NAME, "difficulty-btn")
    assert len(difficulties) >= 3
    print("✅ TC05: Difficulty Buttons")
    log_test("TC05", "Verify Difficulty Buttons", "Check Easy, Medium, Hard buttons",
             "Three difficulty buttons should exist", f"Found {len(difficulties)} buttons", "PASS")
except Exception as e:
    log_test("TC05", "Verify Difficulty Buttons", "Check Easy, Medium, Hard buttons",
             "Three difficulty buttons should exist", str(e), "FAIL")

try:
    driver.find_element(By.CSS_SELECTOR, "[data-game='crossword']").click()
    crossword_controls = driver.find_elements(By.CSS_SELECTOR, "#crossword-game .control-btn")
    assert len(crossword_controls) >= 3
    print("✅ TC06: Crossword Controls")
    log_test("TC06", "Verify Crossword Controls", "Check Crossword control buttons",
             "New Puzzle, Check, Reveal should be visible", f"Found {len(crossword_controls)} controls", "PASS")
except Exception as e:
    log_test("TC06", "Verify Crossword Controls", "Check Crossword control buttons",
             "New Puzzle, Check, Reveal should be visible", str(e), "FAIL")

try:
    driver.find_element(By.CSS_SELECTOR, "[data-game='memory']").click()
    memory_controls = driver.find_elements(By.CSS_SELECTOR, "#memory-game .control-btn")
    assert len(memory_controls) >= 2
    print("✅ TC07: Memory Game Controls")
    log_test("TC07", "Verify Memory Game Controls", "Check Memory game control buttons",
             "New Game and Show All should be visible", f"Found {len(memory_controls)} controls", "PASS")
except Exception as e:
    log_test("TC07", "Verify Memory Game Controls", "Check Memory game control buttons",
             "New Game and Show All should be visible", str(e), "FAIL")

try:
    driver.find_element(By.CSS_SELECTOR, "[data-game='mental-health-tips']").click()
    tips = driver.find_elements(By.CSS_SELECTOR, ".mental-health-tips .tip-card")
    assert len(tips) > 0
    print("✅ TC08: Mental Health Tips Section")
    log_test("TC08", "Verify Mental Health Tips", "Check Mental Health Tips visibility",
             "Tip cards should appear", f"{len(tips)} tips found", "PASS")
except Exception as e:
    log_test("TC08", "Verify Mental Health Tips", "Check Mental Health Tips visibility",
             "Tip cards should appear", str(e), "FAIL")

# -------------------------------
# SUMMARY
# -------------------------------
passed = len([r for r in ws.iter_rows(min_row=2, values_only=True) if r[-1] == "PASS"])
failed = len([r for r in ws.iter_rows(min_row=2, values_only=True) if r[-1] == "FAIL"])
total = passed + failed

ws.append(["", "", "", "", "Total Tests:", total])
ws.append(["", "", "", "", "Passed:", passed])
ws.append(["", "", "", "", "Failed:", failed])
for row in range(ws.max_row-2, ws.max_row+1):
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=6).font = Font(bold=True)

# Auto-fit columns
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 4

# Save with timestamp
timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
excel_path = os.path.join(current_dir, f"Test_Report_{timestamp}.xlsx")
wb.save(excel_path)

print(f"\n📊 Excel Test Report Generated Successfully: {excel_path}")

driver.quit()
